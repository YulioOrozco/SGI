from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Materiales, Usuario, Clientes, PrestamosConsumibles, PrestamosDevolutivos
from .choices import *
from django.http import FileResponse
from django.template import loader
from django.conf import settings
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.core.mail import send_mail
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


# Create your views here.

def generar_excel_materiales(request):
    # Obtén los datos de los materiales desde el modelo
    materiales = Materiales.objects.all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('E1:K1')
    ws.merge_cells('E2:K2')
    ws.merge_cells('E3:K3')

    # Aplica estilos al rango combinado para título
    titulo_celda = ws['E1']
    titulo_celda.value = 'Reporte de Materiales (General)'
    titulo_celda.font = Font(size=22, color="39A900", bold=True)
    titulo_celda.alignment = Alignment(horizontal='center')

    # Aplica estilos al rango combinado para fecha
    fecha_celda = ws['E2']
    fecha_celda.value = 'Fecha: {}'.format(datetime.now().strftime('%Y-%m-%d'))
    fecha_celda.font = Font(size=12, bold=True)
    fecha_celda.alignment = Alignment(horizontal='center')

    # Aplica estilos al rango combinado para nombre de usuario
    nombre_user_celda = ws['E3']
    nombre_user_celda.value = 'Generado por: {}'.format(request.user.username)
    nombre_user_celda.font = Font(size=12, bold=True)
    nombre_user_celda.alignment = Alignment(horizontal='center')

    ws.insert_rows(4)

    
    
    img = Image('App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG')
    img.anchor = 'Q4'
    ws.add_image(img)


    estado_material_map = {
    'Dis': 'Disponible',
    'Pres': 'Préstamo',
    'Gara': 'Garantía',
    'Sop': 'Soporte',
    'DB': 'De baja',
    'Entr': 'Entregado',
    }

    ubicacion_material_map = {
    'Bod': 'Bodega',
    'Z1': 'Zona 1',
    'Z2': 'Zona 2',
    'Z3': 'Zona 3',
    'Z4': 'Zona 4',
    'Z5': 'Zona 5',
    'Z6': 'Zona 6',
    'Admin': 'Administrativos',
    'N.A': 'No aplica',
    'Comp': 'Competencia',
    }



    # Agrega encabezados al archivo Excel
    encabezados = ["ID", "Tipo Material", "Nombre Material", "Modelo Material", "Ubicación Material", "Valor Material", "Estado Material", "Especificación Técnica", "Instructor Encargado", "Código Barras Original", "Código Barras Sena", "Encargado Registrar", "Fecha Ingreso", "Actualización"]
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        celda = '{}5'.format(col_letra)
        ws[celda] = encabezado
        ws[celda].font = Font(size=13, color="39A900", bold=True)
        ws[celda].alignment = Alignment(horizontal='center')

    for idx, material in enumerate(materiales, start=6):
        ws['A{}'.format(idx)] = material.id
        ws['A{}'.format(idx)].font = Font(color="39A900", size=11)

        ws['B{}'.format(idx)] = material.get_tipo_material_display()
        ws['B{}'.format(idx)].font = Font(color="000000", size=11)

        ws['C{}'.format(idx)] = material.nombre_material
        ws['C{}'.format(idx)].font = Font(color="000000", size=11)

        ws['D{}'.format(idx)] = material.modelo_material
        ws['D{}'.format(idx)].font = Font(color="000000", size=11)

        ws['E{}'.format(idx)] = ubicacion_material_map.get(material.ubicacion_material, material.ubicacion_material)
        ws['E{}'.format(idx)].font = Font(color="000000", size=11)

        ws['F{}'.format(idx)] = material.valor_material
        ws['F{}'.format(idx)].font = Font(color="000000", size=11)

        ws['G{}'.format(idx)] = estado_material_map.get(material.estado_material, material.estado_material)
        ws['G{}'.format(idx)].font = Font(color="000000", size=11)

        ws['H{}'.format(idx)] = material.especificacion_tecnica_material
        ws['H{}'.format(idx)].font = Font(color="000000", size=11)

        ws['I{}'.format(idx)] = material.instructor_ecargado_material.nombre_completo() 
        ws['I{}'.format(idx)].font = Font(color="000000", size=11)

        ws['J{}'.format(idx)] = material.codigo_barras_original_material
        ws['J{}'.format(idx)].font = Font(color="000000", size=11)

        ws['K{}'.format(idx)] = material.codigo_barras_sena_material
        ws['K{}'.format(idx)].font = Font(color="000000", size=11)

        ws['L{}'.format(idx)] = material.encargado_registrar_material.nombre_completo()  
        ws['L{}'.format(idx)].font = Font(color="000000", size=11)

        ws['M{}'.format(idx)] = material.fecha_ingreso_material
        ws['M{}'.format(idx)].font = Font(color="000000", size=11)

        ws['N{}'.format(idx)] = material.actualizacion_material
        ws['N{}'.format(idx)].font = Font(color="000000", size=11)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="39A900"),
                                 right=Side(style='thin', color="39A900"),
                                 top=Side(style='thin', color="39A900"),
                                 bottom=Side(style='thin', color="39A900"))

    # Después de ingresar los datos en las celdas, itera sobre las columnas y ajusta el ancho automáticamente
    for column_idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajusta el ancho de la columna para dejar espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Responde con el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_material.xlsx'
    wb.save(response)
    return response

    return render(request, 'App_SistemaGestionInventario/instructor_planta/generar_reporte.html')



def generar_reporte_pdf_materiales(request):

    #usuario = request.user
    

    tipo_material_seleccionado = request.POST.get('tipo_elemento', 'opcion1')
    tipo_material = ''

     # Filtrar materiales según la selección del usuario
    if tipo_material_seleccionado == 'opcion1':  # Todos los materiales
        datos = Materiales.objects.all()
        tipo_material = "Optimización de Inventario"
    elif tipo_material_seleccionado == 'opcion2':  # Consumible
        datos = Materiales.objects.filter(tipo_material='Consumible')
        tipo_material = "Seguimiento del Uso y Desperdicio"
    elif tipo_material_seleccionado == 'opcion3':  # Devolutivo
        datos = Materiales.objects.filter(tipo_material='Devolutivo')
        tipo_material = "Análisis de Tendencias de Consumo"
    elif tipo_material_seleccionado == 'opcion4':  # Devolutivo
        datos = Materiales.objects.filter(tipo_material='Devolutivo')
        tipo_material = "Otra"


    if request.method == 'POST':
        tipo_elemento = request.POST.get('tipo_elemento', 'opcion1')
        periodicidad = request.POST.get('periodicidad', 'opcion1')
        #fecha_inicio = request.POST.get('fecha_inicio', (datetime.now().replace(microsecond=0) - timedelta(days=7)))
        fecha_final = request.POST.get('fecha_final', datetime.now().strftime('%Y-%m-%d'))

        # Lógica para obtener los datos según los parámetros del formulario
        if periodicidad == 'opcion1':  # Dia
            fecha_inicio = datetime.now() - timedelta(days=1)
        elif periodicidad == 'opcion2':  # Semanal
            fecha_inicio = datetime.now() - timedelta(days=7)            
        elif periodicidad == 'opcion3':  # Mensual
            fecha_inicio = datetime.now() - timedelta(days=30)
        elif periodicidad == 'opcion4':  # Trimestral
            fecha_inicio = datetime.now() - timedelta(days=90)
        elif periodicidad == 'opcion5':  # Semestral
            fecha_inicio = datetime.now() - timedelta(days=180)
        elif periodicidad == 'opcion6':  # Anual
            fecha_inicio = datetime.now() - timedelta(days=365)
        elif periodicidad == 'opcion7':  # Bianual
            fecha_inicio = datetime.now() - timedelta(days=730)                                    
        elif periodicidad == 'opcion8':  # Trienal
            fecha_inicio = datetime.now() - timedelta(days=1095)
        elif periodicidad == 'opcion9':  # Cuatrienal
            fecha_inicio = datetime.now() - timedelta(days=1460)
        elif periodicidad == 'opcion10':  # Quinquenal
            fecha_inicio = datetime.now() - timedelta(days=1825)
        elif periodicidad == 'opcion11':  # Todo
            fecha_inicio = datetime.now() - timedelta(days=3650)          
        datos = Materiales.objects.filter(fecha_ingreso_material__gte=fecha_inicio)

        response = BytesIO()
        p = canvas.Canvas(response, pagesize=letter)    

        ancho_pagina = 612  # Tamaño de la página letter en puntos
        numero_de_columnas = 9  # Número total de columnas
        ancho_columna = ancho_pagina / numero_de_columnas
        alto_celda = 10  # Tamaño del texto de la celda
        espacio_entre_columnas = -1  # Espacio entre las columnas

        x = 1
        y = 415  # Ajusta la posición vertical de la tabla según tus necesidades

        # Encabezados de la tabla
        p.setFont("Helvetica-Bold", 6)  # Tamaño de la fuente para los encabezados
        p.setFillColor(colors.HexColor("#39A900"))  # Color verde para los encabezados
        p.drawString(x + 30 + espacio_entre_columnas, y,"ID")
        p.drawString(x + ancho_columna + espacio_entre_columnas, y, "Tipo")
        p.drawString(x + 2 * (ancho_columna + espacio_entre_columnas), y, "Nombre")
        p.drawString(x + 3 * (ancho_columna + espacio_entre_columnas), y, "Estado")
        p.drawString(x + 4 * (ancho_columna + espacio_entre_columnas), y, "Ubicación")
        p.drawString(x + 5 * (ancho_columna + espacio_entre_columnas), y, "Modelo")
        p.drawString(x + 6 * (ancho_columna + espacio_entre_columnas), y, "Responsable")
        p.drawString(x + 7 * (ancho_columna + espacio_entre_columnas), y, "Código Barras Sena")
        p.drawString(x + 8 * (ancho_columna + espacio_entre_columnas), y, "Código Barras Original")

        y -= alto_celda  # Moverse hacia arriba para los datos

        # Agregar los datos a la tabla
        for dato in datos:
            tipo_material_descripcion = ''
            estado_material_descripcion = ''
            ubicacion_material_descripcion = ''

            if dato.tipo_material == 'Consu':
                tipo_material_descripcion = 'Consumible'
            elif dato.tipo_material == 'Devo':
                tipo_material_descripcion = 'Devolutivo'
            # Agrega más condiciones para otros tipos de materiales si es necesario

            if dato.estado_material == 'Dis':
                estado_material_descripcion = 'Disponible'
            elif dato.estado_material == 'Pres':
                estado_material_descripcion = 'Prestamo'
            elif dato.estado_material == 'Gara':
                estado_material_descripcion = 'Garantia'
            elif dato.estado_material == 'Sop':
                estado_material_descripcion = 'Soporte'
            elif dato.estado_material == 'DB':
                estado_material_descripcion = 'De baja'
            elif dato.estado_material == 'Entr':
                estado_material_descripcion = 'Entregado'


            if dato.ubicacion_material == 'Bod':
                ubicacion_material_descripcion = 'Bodega'
            elif dato.ubicacion_material == 'Z1':
                ubicacion_material_descripcion = 'Zona 1'
            elif dato.ubicacion_material == 'Z2':
                ubicacion_material_descripcion = 'Zona 2'
            elif dato.ubicacion_material == 'Z3':
                ubicacion_material_descripcion = 'Zona 3'
            elif dato.ubicacion_material == 'Z4':
                ubicacion_material_descripcion = 'Zona 4'
            elif dato.ubicacion_material == 'Z5':
                ubicacion_material_descripcion = 'Zona 5'
            elif dato.ubicacion_material == 'Z6':
                ubicacion_material_descripcion = 'Zona 6'
            elif dato.ubicacion_material == 'Admin':
                ubicacion_material_descripcion = 'Administrativos'
            elif dato.ubicacion_material == 'N.A':
                ubicacion_material_descripcion = 'No aplica'                
            elif dato.ubicacion_material == 'Comp':
                ubicacion_material_descripcion = 'Competencia' 


            p.setFont("Helvetica", 5)  # Tamaño de la fuente para los datos
            p.setFillColor(colors.HexColor("#050000"))  # Color negro para los datos de la tabla
            p.drawString(x + ancho_columna + espacio_entre_columnas, y, tipo_material_descripcion)
            p.drawString(x + 3 * (ancho_columna + espacio_entre_columnas), y, estado_material_descripcion)
            p.drawString(x + 4 * (ancho_columna + espacio_entre_columnas), y, ubicacion_material_descripcion)
            p.drawString(x + 30 + espacio_entre_columnas, y, str(dato.id))        
            p.drawString(x + 2 * (ancho_columna + espacio_entre_columnas), y, dato.nombre_material)        
            p.drawString(x + 5 * (ancho_columna + espacio_entre_columnas), y, dato.modelo_material)
            
            # Aquí obtenemos el nombre completo del instructor como una cadena
            if dato.instructor_ecargado_material:
                nombre_instructor = dato.instructor_ecargado_material.nombre_completo()
                p.drawString(x + 6 * (ancho_columna + espacio_entre_columnas), y, nombre_instructor)
            else:
                p.drawString(x + 6 * (ancho_columna + espacio_entre_columnas), y, "Sin instructor asignado")
            
            p.drawString(x + 7 * (ancho_columna + espacio_entre_columnas), y, str(dato.codigo_barras_sena_material))
            p.drawString(x + 8 * (ancho_columna + espacio_entre_columnas), y, str(dato.codigo_barras_original_material))
            
            y -= alto_celda  # Mueve la posición Y para el próximo dato


        #usuario = User.objects.get(id=id_del_usuario)

        # Imagen en una posición específica
        image_path = "App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG"
        image_x = 255  # posición x de la imagen
        image_y = letter[1] - 110  # posición y de la imagen
        p.drawImage(image_path, image_x, image_y, width=1.2*inch, height=1.2*inch)

        # Agregar texto en posiciones específicas
        titulo = "Centro De Diseño E Innovacion Tecnologica Industrial"
        descripcion = "Dg. 27a #4-2 a 4-114, Dosquebradas, Risaralda"
        otra_descripcion = "Sistema de Gestión de Inventario"

        titulo_reporte = "Reporte de Materiales"
        fecha_reporte = "Generado el "
        muestra_fecha_reporte = "{}".format(datetime.now().strftime("%Y-%m-%d %H:%M"))
        autor_reporte = "El autor del reporte:" 
        reporte_autor = f"Francis Quizcer" #{usuario.username}
        tippo_material = f"Razón:"
        muestra_tipo_material = f"{tipo_material}"
        tiempo = f"Tiempo Seleccionado:"
        muestra_tiempo = f"Desde {fecha_inicio} hasta {fecha_final}"  

        p.setFont("Helvetica-Bold", 14)
        color_verde = colors.HexColor("#39A900")
        p.setFillColor(color_verde)         
        p.drawString(1.8*inch, letter[1] - 2.1*inch, titulo)
        p.drawString(3*inch, letter[1] - 1.8*inch, otra_descripcion)
        color_negro = colors.HexColor("#050000")
        p.setFillColor(color_negro)     
        p.setFont("Helvetica", 12)
        p.drawString(2.5*inch, letter[1] - 2.4*inch, descripcion)
        
        p.setFont("Helvetica-Bold", 14)
        color_verde = colors.HexColor("#39A900")
        p.setFillColor(color_verde) 
        p.drawString(3.2*inch, letter[1] - 3.3*inch, titulo_reporte)
        p.setFont("Helvetica", 12)
        color_negro = colors.HexColor("#050000")
        p.setFillColor(color_negro)     
        p.drawString(1*inch, letter[1] - 4*inch, fecha_reporte)
        p.drawString(1*inch, letter[1] - 4.3*inch, autor_reporte)
        p.drawString(5.0*inch, letter[1] - 4.0*inch, tippo_material)
        p.drawString(5.0*inch, letter[1] - 4.3*inch, tiempo )
        p.setFillColor(color_verde) 
        p.setFont("Helvetica-Bold", 10)
        p.drawString(4.6*inch, letter[1] - 4.6*inch, muestra_tiempo )
        p.drawString(1.3*inch, letter[1] - 4.6*inch, reporte_autor )
        p.drawString(5.7*inch, letter[1] - 4.0*inch, muestra_tipo_material )
        p.drawString(2.0*inch, letter[1] - 4.0*inch, muestra_fecha_reporte )

        #y = 700

        #espacio_restante = y

        # Si no hay suficiente espacio en la primera página, agregar una nueva página
        #if espacio_restante < 50:  # Ajusta este valor según sea necesario
            #p.showPage()  # Agregar una nueva página
            #y = 700  # Posición vertical inicial en la segunda página        
        
        p.showPage()
        p.save()

        response.seek(0)        

        # Devolver el PDF como respuesta para su descarga
        return FileResponse(response, content_type='application/pdf')

    return render(request, 'App_SistemaGestionInventario/instructor_planta/generar_reporte.html')



# Create your views here.

#vista PDF

def generar_excel_prestamo(request):
    # Obtén los datos de los materiales desde el modelo
    prestamos_devolutivos = PrestamosDevolutivos.objects.select_related('encargado_registra_material_devolutivo',
                                                                        'recibe_prestamo_material_devolutivo',
                                                                        'material_otorgado_devolutivo').all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active

    img = Image('App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG')
    img.anchor = 'Q4'
    ws.add_image(img)

    # Aplica estilos al rango combinado para título, fecha y nombre de usuario
    ws.merge_cells('E1:K1')
    ws.merge_cells('E2:K2')
    ws.merge_cells('E3:K3')

    titulo_celda = ws['E1']
    titulo_celda.value = 'Reporte de Materiales Devolutivos Prestados'
    titulo_celda.font = Font(size=22, color="39A900", bold=True)
    titulo_celda.alignment = Alignment(horizontal='center')

    fecha_celda = ws['E2']
    fecha_celda.value = 'Fecha: {}'.format(datetime.now().strftime('%Y-%m-%d'))
    fecha_celda.font = Font(size=12, bold=True)
    fecha_celda.alignment = Alignment(horizontal='center')

    nombre_user_celda = ws['E3']
    nombre_user_celda.value = 'Generado por: {}'.format(request.user.username)
    nombre_user_celda.font = Font(size=12, bold=True)
    nombre_user_celda.alignment = Alignment(horizontal='center')

    # Agrega encabezados al archivo Excel
    encabezados = ["ID", "Encargado Registra", "Quien recibe", "ID Material",
                   "Tipo Material", "Nombre Material", "Modelo Material", "Ubicación Material", "Valor Material",
                   "Fecha Entrega", "Fecha Devolución"]
   
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        celda = '{}5'.format(col_letra)
        ws[celda] = encabezado
        ws[celda].font = Font(size=13, color="39A900", bold=True)
        ws[celda].alignment = Alignment(horizontal='center')

    # Llena el archivo Excel con los datos de los préstamos devolutivos
    for idx, prestamo in enumerate(prestamos_devolutivos, start=6):
        ws['A{}'.format(idx)] = prestamo.id
        ws['B{}'.format(idx)] = prestamo.encargado_registra_material_devolutivo.nombre_completo()        
        ws['C{}'.format(idx)] = prestamo.recibe_prestamo_material_devolutivo.nombre_cliente()        
        ws['D{}'.format(idx)] = prestamo.material_otorgado_devolutivo.id
        ws['E{}'.format(idx)] = prestamo.material_otorgado_devolutivo.get_tipo_material_display()
        ws['F{}'.format(idx)] = prestamo.material_otorgado_devolutivo.nombre_material
        ws['G{}'.format(idx)] = prestamo.material_otorgado_devolutivo.modelo_material
        ws['H{}'.format(idx)] = prestamo.material_otorgado_devolutivo.ubicacion_material
        ws['I{}'.format(idx)] = prestamo.material_otorgado_devolutivo.valor_material
        ws['J{}'.format(idx)] = prestamo.fecha_entrega_material_devolutivo.strftime('%Y-%m-%d')
        ws['K{}'.format(idx)] = prestamo.fecha_devolucion_material_devolutivo.strftime('%Y-%m-%d')

    # Aplica bordes a las celdas
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="39A900"),
                                 right=Side(style='thin', color="39A900"),
                                 top=Side(style='thin', color="39A900"),
                                 bottom=Side(style='thin', color="39A900"))

    # Ajusta el ancho de las columnas
    for column_idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajusta el ancho de la columna para dejar espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Responde con el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_prestamo_devolutivo.xlsx'
    wb.save(response)
    return response



def generar_excel_consumible(request):
    # Obtén los datos de los materiales desde el modelo
    prestamos_consumible = PrestamosConsumibles.objects.select_related('encargado_registra_prestamo_consumible',
                                                                        'recibe_prestamo_prestamo_consumible',
                                                                        'material_otorgado_prestamo_consumible').all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active

    img = Image('App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG')
    img.anchor = 'L4'
    ws.add_image(img)

    # Aplica estilos al rango combinado para título, fecha y nombre de usuario
    ws.merge_cells('C1:E1')
    ws.merge_cells('C2:E2')
    ws.merge_cells('C3:E3')

    titulo_celda = ws['C1']
    titulo_celda.value = 'Reporte de Materiales Consumibles Otorgados'
    titulo_celda.font = Font(size=22, color="39A900", bold=True)
    titulo_celda.alignment = Alignment(horizontal='center')

    fecha_celda = ws['C2']
    fecha_celda.value = 'Fecha: {}'.format(datetime.now().strftime('%Y-%m-%d'))
    fecha_celda.font = Font(size=12, bold=True)
    fecha_celda.alignment = Alignment(horizontal='center')

    nombre_user_celda = ws['C3']
    nombre_user_celda.value = 'Generado por: {}'.format(request.user.username)
    nombre_user_celda.font = Font(size=12, bold=True)
    nombre_user_celda.alignment = Alignment(horizontal='center')

    # Agrega encabezados al archivo Excel
    encabezados = ["ID", "Encargado Registra", "Quien recibe", "Ubicación Material",
                   "Nombre Material","Fecha Entrega"]
   
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        celda = '{}5'.format(col_letra)
        ws[celda] = encabezado
        ws[celda].font = Font(size=13, color="39A900", bold=True)
        ws[celda].alignment = Alignment(horizontal='center')

    # Llena el archivo Excel con los datos de los préstamos devolutivos
    for idx, prestamo in enumerate(prestamos_consumible, start=6):
        ws['A{}'.format(idx)] = prestamo.id
        ws['B{}'.format(idx)] = prestamo.encargado_registra_prestamo_consumible.nombre_completo()        
        ws['C{}'.format(idx)] = prestamo.recibe_prestamo_prestamo_consumible.nombre_cliente()        
        ws['D{}'.format(idx)] = prestamo.ubicacion_prestamo_prestamo_consumible
        ws['E{}'.format(idx)] = prestamo.material_otorgado_prestamo_consumible.get_tipo_material_display()      
        ws['F{}'.format(idx)] = prestamo.fecha_entrega_prestamo_consumible.strftime('%Y-%m-%d')
        

    # Aplica bordes a las celdas
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="39A900"),
                                 right=Side(style='thin', color="39A900"),
                                 top=Side(style='thin', color="39A900"),
                                 bottom=Side(style='thin', color="39A900"))

    # Ajusta el ancho de las columnas
    for column_idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajusta el ancho de la columna para dejar espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Responde con el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_prestamo_consumible.xlsx'
    wb.save(response)
    return response



def generar_excel_clientes(request):
    # Obtén los datos de los préstamos devolutivos desde el modelo
    reporte_clientes = Clientes.objects.select_related().all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active


    img = Image('App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG')
    img.anchor = 'S4'
    ws.add_image(img)

    # Aplica estilos al rango combinado para título, fecha y nombre de usuario
    ws.merge_cells('E1:K1')
    ws.merge_cells('E2:K2')
    ws.merge_cells('E3:K3')

    titulo_celda = ws['E1']
    titulo_celda.value = 'Reporte de Internautas'
    titulo_celda.font = Font(size=22, color="39A900", bold=True)
    titulo_celda.alignment = Alignment(horizontal='center')

    fecha_celda = ws['E2']
    fecha_celda.value = 'Fecha: {}'.format(datetime.now().strftime('%Y-%m-%d'))
    fecha_celda.font = Font(size=12, bold=True)
    fecha_celda.alignment = Alignment(horizontal='center')

    nombre_user_celda = ws['E3']
    nombre_user_celda.value = 'Generado por: {}'.format(request.user.username)
    nombre_user_celda.font = Font(size=12, bold=True)
    nombre_user_celda.alignment = Alignment(horizontal='center')

    # Agrega encabezados al archivo Excel
    encabezados = ["ID", "Rol", "Tipo Documento", "Numero Documento",
                   "Primer Nombre", "Segundo Nombre", "Primer Apellido",
                   "Segundo Apellido", "Correo soysena", "Primer Telefono",
                   "Segundo Telefono", "Numero Ficha", "Fecha Ingreso Sistema"]

    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        celda = '{}5'.format(col_letra)
        ws[celda] = encabezado
        ws[celda].font = Font(size=13, color="39A900", bold=True)
        ws[celda].alignment = Alignment(horizontal='center')

    # Llena el archivo Excel con los datos de los préstamos devolutivos
    for idx, prestamo in enumerate(reporte_clientes, start=6):
        ws['A{}'.format(idx)] = prestamo.id
        ws['B{}'.format(idx)] = prestamo.rol
        ws['C{}'.format(idx)] = prestamo.tipo_documento
        ws['D{}'.format(idx)] = prestamo.numero_documento
        ws['E{}'.format(idx)] = prestamo.primer_nombre
        ws['F{}'.format(idx)] = prestamo.segundo_nombre
        ws['G{}'.format(idx)] = prestamo.primer_apellido
        ws['H{}'.format(idx)] = prestamo.segundo_apellido
        ws['I{}'.format(idx)] = prestamo.correo_soy_sena
        ws['J{}'.format(idx)] = prestamo.primer_telefono
        ws['K{}'.format(idx)] = prestamo.segundo_telefono
        ws['L{}'.format(idx)] = prestamo.numero_ficha
        ws['M{}'.format(idx)] = prestamo.fecha_ingres_sistema.strftime('%Y-%m-%d')


    # Aplica bordes a las celdas
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="39A900"),
                                 right=Side(style='thin', color="39A900"),
                                 top=Side(style='thin', color="39A900"),
                                 bottom=Side(style='thin', color="39A900"))

    # Ajusta el ancho de las columnas
    for column_idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajusta el ancho de la columna para dejar espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Responde con el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_internauta.xlsx'
    wb.save(response)
    return response



def generar_excel_usuario(request):
    # Obtén los datos de los materiales desde el modelo
    reporte_usuario = Usuario.objects.select_related().all()

    # Crea un nuevo libro de Excel y selecciona la hoja activa
    wb = Workbook()
    ws = wb.active

    img = Image('App_SistemaGestionInventario/static/App_SistemaGestionInventario/img/imasenapdf.PNG')
    img.anchor = 'V4'
    ws.add_image(img)

    # Aplica estilos al rango combinado para título, fecha y nombre de usuario
    ws.merge_cells('E1:K1')
    ws.merge_cells('E2:K2')
    ws.merge_cells('E3:K3')

    titulo_celda = ws['E1']
    titulo_celda.value = 'Reporte de Usuarios'
    titulo_celda.font = Font(size=22, color="39A900", bold=True)
    titulo_celda.alignment = Alignment(horizontal='center')

    fecha_celda = ws['E2']
    fecha_celda.value = 'Fecha: {}'.format(datetime.now().strftime('%Y-%m-%d'))
    fecha_celda.font = Font(size=12, bold=True)
    fecha_celda.alignment = Alignment(horizontal='center')

    nombre_user_celda = ws['E3']
    nombre_user_celda.value = 'Generado por: {}'.format(request.user.username)
    nombre_user_celda.font = Font(size=12, bold=True)
    nombre_user_celda.alignment = Alignment(horizontal='center')

    # Agrega encabezados al archivo Excel
    encabezados = ["ID", "Nombre Uno", "Nombre Dos", "Apellido Uno", "Apellido Dos",
                   "Dd Tipo Documento", "Numero Documento", "Correo Sena", "Correo soysena", "Celular 1",
                   "Celular 2", "Id Rol", "Fecha Inicio Contrato", "Fecha Fin Contrato", "Id Area Instrtuctor", "Estado Cuenta"]
   
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        celda = '{}5'.format(col_letra)
        ws[celda] = encabezado
        ws[celda].font = Font(size=13, color="39A900", bold=True)
        ws[celda].alignment = Alignment(horizontal='center')

    # Llena el archivo Excel con los datos de los préstamos devolutivos
    for idx, prestamo in enumerate(reporte_usuario, start=6):
        ws['A{}'.format(idx)] = prestamo.id
        ws['B{}'.format(idx)] = prestamo.nombre_1
        ws['C{}'.format(idx)] = prestamo.nombre_2
        ws['D{}'.format(idx)] = prestamo.apellido_1
        ws['E{}'.format(idx)] = prestamo.apellido_2
        ws['F{}'.format(idx)] = prestamo.id_tipo_documento
        ws['G{}'.format(idx)] = prestamo.numero_documento
        ws['H{}'.format(idx)] = prestamo.correo_sena
        ws['I{}'.format(idx)] = prestamo.correo_soy_sena
        ws['J{}'.format(idx)] = prestamo.celular_1
        ws['K{}'.format(idx)] = prestamo.celular_2
        ws['L{}'.format(idx)] = prestamo.id_rol
        ws['M{}'.format(idx)] = prestamo.fecha_inicio_contrato.strftime('%Y-%m-%d')
        ws['N{}'.format(idx)] = prestamo.fecha_fin_contrato.strftime('%Y-%m-%d')
        ws['O{}'.format(idx)] = prestamo.id_area_instrtuctor
        ws['P{}'.format(idx)] = prestamo.estado_cuenta
        

    # Aplica bordes a las celdas
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color="39A900"),
                                 right=Side(style='thin', color="39A900"),
                                 top=Side(style='thin', color="39A900"),
                                 bottom=Side(style='thin', color="39A900"))

    # Ajusta el ancho de las columnas
    for column_idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajusta el ancho de la columna para dejar espacio adicional
        ws.column_dimensions[column_letter].width = adjusted_width

    # Responde con el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_usuario.xlsx'
    wb.save(response)
    return response






def enviar_recordatorio_correo(usuario, cliente, material, fecha_entrega_material_devolutivo, fecha_devolucion_material_devolutivo):
    # Construir el asunto del correo
    asunto = 'Recordatorio de Entrega de Material Devolutivo'

    # Construir el mensaje del correo
    mensaje = f'Estimad@ {cliente.primer_nombre} {cliente.primer_apellido},\n\n' \
              f'Este es un recordatorio de que debe entregar el material devolutivo "{material.nombre_material}" ' \
              f'a {usuario.nombre_1} {usuario.apellido_1} suministrado el {fecha_entrega_material_devolutivo}.\n' \
              f'La fecha límite para la devolución es el {fecha_devolucion_material_devolutivo}.\n' \
              f'Por favor, asegúrese de hacer la entrega a tiempo.\n\n' \
              f'Atentamente,\n' \
              f'Sistema de Gestión de Inventario - CDITI'

    # Dirección de correo electrónico del remitente
    remitente = settings.EMAIL_HOST_USER

    # Direcciones de correo electrónico de los destinatarios (puedes pasar una lista de correos electrónicos aquí)
    destinatarios = [usuario.correo_soy_sena, cliente.correo_soy_sena]

    # Envía el correo electrónico
    send_mail(asunto, mensaje, remitente, destinatarios, fail_silently=False)



def generar_reporte(request):
    return render (request, "App_SistemaGestionInventario/instructor_planta/generar_reporte.html")



#vistas generales para los roles

def home(request):
    return render (request, "App_SistemaGestionInventario/general/index_principal.html")

def index_info(request):
    return render (request, "App_SistemaGestionInventario/general/index_info.html")

def login_cuenta(request):
    return render (request, "App_SistemaGestionInventario/general/login.html")

def verificar_usuario(request):
    return render (request, "App_SistemaGestionInventario/general/verificar_numero_usuario.html")

def verificar_contrasena(request):
    return render (request, "App_SistemaGestionInventario/general/verificar_contrasena.html")



#-----------------------------   FUNCIONES   -----------------------------

def campos_llenos(formulario):
    for campos in formulario:
        if campos == None or campos == '':
            return False
    return True



def verificar_codigo_sena(codigo_sena):
    try:
        db = Materiales.objects.get(codigo_barras_sena_material = codigo_sena)
        return False
    except Materiales.DoesNotExist:
        return True



def traer_objeto_usuario(numero_instructor):
    instructor_encargado = Usuario.objects.get(numero_documento = numero_instructor)
    return instructor_encargado



def conversion_tipo_material(tipo_elemento):
    for tipo in tipo_material:
        if tipo[1] == tipo_elemento:
            tipo_elemento = tipo[0]
            break
    return tipo_elemento



def conversion_ubicacion_material(ubicacion_elemento):
    for ubicacion in ubicacion_material:
        if ubicacion[1] == ubicacion_elemento:
            ubicacion_elemento = ubicacion[0]
    return ubicacion_elemento



def conversion_estado_material(estado_elemento):
    for estado in estado_material:
        if estado[1] == estado_elemento:
            estado_elemento = estado[0]
    return estado_elemento



def conversion_estado_listar_material(materiales):
    for material in materiales:
        if material.estado_material == estado_material[0][0]:
            material.estado_material = estado_material[0][1]

        elif material.estado_material == estado_material[1][0]:
            material.estado_material = estado_material[1][1]
        
        elif material.estado_material == estado_material[2][0]:
            material.estado_material = estado_material[2][1]

        elif material.estado_material == estado_material[3][0]:
            material.estado_material = estado_material[3][1]

        elif material.estado_material == estado_material[4][0]:
            material.estado_material = estado_material[4][1]

        elif material.estado_material == estado_material[5][0]:
            material.estado_material = estado_material[5][1]

    return materiales



def conversion_rol_cuenta(usuarios):
    contador = 0
    for usuario in usuarios:
        if usuario.id_rol == rol[contador][0]:
            usuario.id_rol = rol[contador][1]
        contador += 1
    return usuarios



def conversion_estado_cuenta(usuarios):
    for usuario in usuarios:
        if usuario.estado_cuenta == estado_cuenta_usuario[0][0]:
            usuario.estado_cuenta = estado_cuenta_usuario[0][1]
        elif usuario.estado_cuenta == estado_cuenta_usuario[1][0]:
            usuario.estado_cuenta = estado_cuenta_usuario[1][1]
    return usuarios





def comprobar_usuario(numero_documento_recibe):
    numero_documento_recibe = int(numero_documento_recibe)
    try:
        db = Clientes.objects.get(numero_documento = numero_documento_recibe)
        return True
    except Clientes.DoesNotExist:
        return False



#Vistas para el instructor de planta

def funciones_instructor_planta(request):
    return render (request, "App_SistemaGestionInventario/instructor_planta/funciones_planta.html")






def registrar_material_instructor_planta(request):
    
    #Bloque para desplegables para registrar materiales
    usuarios = Usuario.objects.all()

    #Instructores que se asignan materiales (planta)
    ins_planta = Usuario.objects.filter(id_rol = 'I.P')
    
    lugares = []
    for lugar in ubicacion_material:
        lugares.append(lugar[1])
        
    
    material_estado = []
    for estado in estado_material:
        material_estado.append(estado[1])
        

    material_tipo = []
    for tipo in tipo_material:
        material_tipo.append(tipo[1])


    
    #Bloque para confirmacion e ingreso de nuevos materiales al sistema
    if request.method == 'POST':
        
        #Traer los valores del front
        tipo_elemento = request.POST.get('tipo_material')
        nombre_elemento = request.POST.get('nombre_material')
        modelo_elemento = request.POST.get('modelo_material')
        ubicacion_elemento = request.POST.get('ubicacion_material')
        valor_elemento = request.POST.get('valor_material')
        estado_elemento = request.POST.get('estado_material')
        especificacion_elemento = request.POST.get('especificacion_material')
        instructor_encargado_elemento = request.POST.get('instructor_encargado')
        codigo_barras_normal_elemento = str(request.POST.get('codigo_barras_normal'))
        codigo_sena_elemento = str(request.POST.get('codigo_sena'))
        encargado_registro = request.POST.get('encargado_registro')
        firma_electronica_elemento = request.POST.get('firma_electronica')
        fecha_ingreso_elemento = request.POST.get('fecha_ingreso')
        
        #Se utiliza para guardar las variables que traen los datos de front
        formulario = [
            tipo_elemento,
            nombre_elemento,
            modelo_elemento,
            ubicacion_elemento,
            valor_elemento,
            estado_elemento,
            especificacion_elemento,
            instructor_encargado_elemento,
            codigo_barras_normal_elemento,
            codigo_sena_elemento,
            encargado_registro,
            firma_electronica_elemento,
            fecha_ingreso_elemento,
        ]
        
        mensaje = ''
        
        #Verifica si todas las variables de la lista contienen datos
        if campos_llenos(formulario) == True:

            #Verifica si el codigo SENA que actualmente trae de front se encuentra o no en la db
            if verificar_codigo_sena(codigo_sena_elemento):

                #Llaves foraneas 
                instructor_encargado_elemento = traer_objeto_usuario(instructor_encargado_elemento)
                encargado_registro = traer_objeto_usuario(encargado_registro)


                #Valores de los choices

                #Choices para tipo de material
                tipo_elemento = conversion_tipo_material(tipo_elemento)

                #Choices para ubicacion del material
                ubicacion_elemento = conversion_ubicacion_material(ubicacion_elemento)

                #Choices para estado material
                estado_elemento = conversion_estado_material(estado_elemento)
            
                #Crear variable para guardar todos los datos del modelo "Materiales"
                agregar_material = Materiales(tipo_material = tipo_elemento, nombre_material = nombre_elemento, modelo_material = modelo_elemento, ubicacion_material = ubicacion_elemento, valor_material = valor_elemento, estado_material = estado_elemento, especificacion_tecnica_material = especificacion_elemento, instructor_ecargado_material = instructor_encargado_elemento, codigo_barras_original_material = codigo_barras_normal_elemento, codigo_barras_sena_material = codigo_sena_elemento, encargado_registrar_material = encargado_registro, fecha_ingreso_material = '', actualizacion_material = '')
                
                #Guardar los valores de la variable en el modelo
                agregar_material.save()


                mensaje = 'El material fue registrado.'

            else:
                mensaje = 'El código ya existe, vuelva a intentarlo.'
            
        else:
            mensaje = 'Por favor complete todos los campos.'

        return render (request, "App_SistemaGestionInventario/instructor_planta/registrar_mate_planta.html", {'usuarios':usuarios, 'ins_planta':ins_planta, 'ubicaciones':lugares, 'estado_material':material_estado, 'tipo_material':material_tipo, 'alert_ingre_mate':alerta_ingreso_material, 'alert_campos_llenos':verificar_campos_completos})
        
    return render (request, "App_SistemaGestionInventario/instructor_planta/registrar_mate_planta.html", {'usuarios':usuarios, 'ins_planta':ins_planta, 'ubicaciones':lugares, 'estado_material':material_estado, 'tipo_material':material_tipo, 'alert_ingre_mate':alerta_ingreso_material, 'alert_campos_llenos':verificar_campos_completos})



def lista_usuarios_planta(request):
    ins_planta = Usuario.objects.filter(id_rol = rol[0][0])
    ins_planta = conversion_rol_cuenta(ins_planta)

    ins_contrato = Usuario.objects.filter(id_rol = rol[1][0])
    ins_contrato = conversion_rol_cuenta(ins_contrato)

    monitor = Usuario.objects.filter(id_rol = rol[2][0])
    monitor = conversion_rol_cuenta(monitor)



    #Cambio para que no aparezca una letra, sino todo (Activo, Inactivo)

    ins_planta = conversion_estado_cuenta(ins_planta)
    ins_contrato = conversion_estado_cuenta(ins_contrato)
    monitor = conversion_estado_cuenta(monitor)



    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_usuarios_planta.html", {'ins_planta':ins_planta, 'ins_contrato':ins_contrato, 'monitor':monitor})



def listar_materiales_consumibles_planta(request):

    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', nombre_material__icontains = busqueda)


        consulta_db = conversion_estado_listar_material(consulta_db)


        materiales_disponibles = []

        for consulta in consulta_db:
            if consulta.estado_material == 'Disponible':
                materiales_disponibles.append(consulta)


        cantidad_modelo = []


        for consulta in materiales_disponibles:
            modelo_actual = consulta.modelo_material
            encontrado = False

            for lista in cantidad_modelo:
                if lista[0] == modelo_actual:
                    lista[1] += 1
                    encontrado = True
                    break

            if not encontrado:
                cantidad_modelo.append([modelo_actual, 1])

        return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_consu_planta.html", {'materiales_consumibles':consulta_db, 'busqueda':busqueda, 'cantidad_modelo':cantidad_modelo})
        
    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_consu_planta.html")



def listar_materiales_devolutivos_planta(request):

    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', nombre_material__icontains = busqueda)



        consulta_db = conversion_estado_listar_material(consulta_db)
        

        materiales_disponibles = []

        for consulta in consulta_db:
            if consulta.estado_material == 'Disponible':
                materiales_disponibles.append(consulta)
        cantidad_modelo = []


        for consulta in materiales_disponibles:
            modelo_actual = consulta.modelo_material
            encontrado = False

            for lista in cantidad_modelo:
                if lista[0] == modelo_actual:
                    lista[1] += 1
                    encontrado = True
                    break

            if not encontrado:
                cantidad_modelo.append([modelo_actual, 1])
        
        #cantidad[] = Materiales.objects.filter(tipo_material = 'Devo', modelo = (Modelo))

        return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_devo_planta.html", {'consulta_db': consulta_db, 'busqueda':busqueda, 'cantidad_modelo':cantidad_modelo})
    
    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_devo_planta.html")







def calendario_planta(request):
    return render (request, "App_SistemaGestionInventario/instructor_planta/calendario_planta.html")



def listar_material_garantia_planta(request):

    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material_garantia')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterialDevolutivo')
            consulta_db = Materiales.objects.filter(estado_material = 'Gara', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterialDevolutivo')
            consulta_db = Materiales.objects.filter(estado_material = 'Gara', nombre_material__icontains = busqueda)

        
        for consulta in consulta_db:
            for estado in estado_material:
                if estado[0] == consulta.estado_material:
                    consulta.estado_material = estado[1]


        return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_garan_planta.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    
    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_garan_planta.html")



def listar_material_baja_planta(request):

    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'DB', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'DB', nombre_material__icontains = busqueda)

        
        consulta_db = conversion_estado_listar_material(consulta_db)

        return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_baja_planta.html", {'consulta_db':consulta_db, 'busqueda':busqueda})

    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_baja_planta.html")



def listar_material_soporte_planta(request):

    if request.method == 'POST':

        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Sop', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Sop', nombre_material__icontains = busqueda)

        
        for consulta in consulta_db:
            for estado in estado_material:
                if estado[0] == consulta.estado_material:
                    consulta.estado_material = estado[1]
            
        return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_sup_planta.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_planta/listar_mate_sup_planta.html")



def entregable_consumible_planta(request):

    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Consu', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])


    #Verificar el metodo por el cual se esta pasando los datos del front
    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_consumible_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_consumible =request.POST.get('fecha_prestamo')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_consumible,
            ubicacion_material_prestamo,
        ]

        #Verifica si los campos de formulario estan llenos (Va a una funcion)
        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("No hay materiales para entregar")
                return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_consu_planta.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 
                
            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_consu_planta.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 


            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                guardar_nuevo_cliente.save()
            

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_consumible = PrestamosConsumibles(encargado_registra_prestamo_consumible = encargado_prestamo,
                                                                        recibe_prestamo_prestamo_consumible = numero_documento_recibe,
                                                                        ubicacion_prestamo_prestamo_consumible = ubicacion_material_prestamo,
                                                                        material_otorgado_prestamo_consumible = material_entregable,
                                                                        fecha_entrega_prestamo_consumible = fecha_otorga_consumible)
            registrar_nuevo_prestamo_consumible.save()

            #Actualizar el estado del material entregado

            material_entregable.estado_material = estado_material[5][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
               
            
                        
        else:
            print("sssssssssssssssssssssssssssssssssssssssssssss")
    
    return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_consu_planta.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion})



def entregable_devolutivo_planta(request):
    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Devo', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])

    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('rol_persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_devolutivo_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_devolutivo = request.POST.get('fecha_prestamo')
        fecha_devolucion_devolutivo = request.POST.get('fecha_devolucion')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_devolutivo,
            ubicacion_material_prestamo,
            fecha_devolucion_devolutivo,
        ]

        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("CAMINO 3")
                return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_devo_planta.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            print("CAMINO 1")

            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                print("CAMINO 4")
                return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_devo_planta.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            
            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                print("CAMINO 5")
                guardar_nuevo_cliente.save()

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_devolutivo = PrestamosDevolutivos(encargado_registra_material_devolutivo = encargado_prestamo,
                                                                        recibe_prestamo_material_devolutivo = numero_documento_recibe,
                                                                        ubicacion_prestamo_material_devolutivo = ubicacion_material_prestamo,
                                                                        material_otorgado_devolutivo = material_entregable,
                                                                        fecha_entrega_material_devolutivo = '',
                                                                        fecha_devolucion_material_devolutivo = fecha_devolucion_devolutivo)
            registrar_nuevo_prestamo_devolutivo.save()

            #Actualizar el estado del material entregado
            material_entregable.estado_material = estado_material[1][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
            print("CAMINO 6")



            fecha_entrega = datetime.strptime(fecha_otorga_devolutivo, '%Y-%m-%d').date()
            fecha_devolucion = datetime.strptime(fecha_devolucion_devolutivo, '%Y-%m-%d').date()
            fecha_actual = datetime.now()

            un_dia_antes = fecha_devolucion - timedelta(days=0)
        
            #Enviar correos electrónicos de recordatorio uno días antes
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)


        else:
            print("CAMINO 2")
    return render (request, "App_SistemaGestionInventario/instructor_planta/entregable_devo_planta.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})



def visualizar_cuenta_planta(request):
    return render (request, "App_SistemaGestionInventario/instructor_planta/ver_perfil_planta.html")




#Vistas para el instructor contratista

def funciones_instructor_contrato(request):
    return render (request, "App_SistemaGestionInventario/instructor_contratista/funciones_contratista.html")

def registrar_materiales_instructor_contrato(request):
    #Bloque para desplegables

    #Traer a todos los usuarios
    usuarios = Usuario.objects.all()

    #Separa a los instructores contratistas
    ins_planta = Usuario.objects.filter(id_rol = 'I.P')

    #Tipo de material
    material = []

    for mate in tipo_material:
        material.append(mate[1])


    #Ubicacion material
    ubicacion = []

    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])


    #Estado material
    estado_para_elemento = []

    for estado in estado_material:
        estado_para_elemento.append(estado[1])


    #Bloque para confirmacion e ingreso de nuevos materiales al sistema
    if request.method == 'POST':
        
        #Traer los valores del front
        tipo_elemento = request.POST.get('tipo_material')
        nombre_elemento = request.POST.get('nombre_material')
        modelo_elemento = request.POST.get('modelo_material')
        ubicacion_elemento = request.POST.get('ubicacion_material')
        valor_elemento = request.POST.get('valor_material')
        estado_elemento = request.POST.get('estado_material')
        especificacion_elemento = request.POST.get('especificacion_material')
        instructor_encargado_elemento = request.POST.get('instructor_encargado')
        codigo_barras_normal_elemento = str(request.POST.get('codigo_barras_normal'))
        codigo_sena_elemento = str(request.POST.get('codigo_sena'))
        encargado_registro = request.POST.get('encargado_registro')
        firma_electronica_elemento = request.POST.get('firma_electronica')
        fecha_ingreso_elemento = request.POST.get('fecha_ingreso')


        #Se utiliza para guardar las variables que traen los datos de front
        formulario = [
            tipo_elemento,
            nombre_elemento,
            modelo_elemento,
            ubicacion_elemento,
            valor_elemento,
            estado_elemento,
            especificacion_elemento,
            instructor_encargado_elemento,
            codigo_barras_normal_elemento,
            codigo_sena_elemento,
            encargado_registro,
            firma_electronica_elemento,
            fecha_ingreso_elemento,
        ]

        if  campos_llenos(formulario) == True:

            if verificar_codigo_sena(codigo_sena_elemento):

                #Llaves foraneas 
                instructor_encargado_elemento = traer_objeto_usuario(instructor_encargado_elemento) 
                encargado_registro = traer_objeto_usuario(encargado_registro)


                #Valores de los choices

                #Choices para tipo de material
                tipo_elemento = conversion_tipo_material(tipo_elemento)

                #Choices para ubicacion del material
                ubicacion_elemento = conversion_ubicacion_material(ubicacion_elemento)

                #Choices para estado material
                estado_elemento = conversion_estado_material(estado_elemento)


                #Crear variable para guardar todos los datos del modelo "Materiales"
                agregar_material = Materiales(tipo_material = tipo_elemento, nombre_material = nombre_elemento, modelo_material = modelo_elemento, ubicacion_material = ubicacion_elemento, valor_material = valor_elemento, estado_material = estado_elemento, especificacion_tecnica_material = especificacion_elemento, instructor_ecargado_material = instructor_encargado_elemento, codigo_barras_original_material = codigo_barras_normal_elemento, codigo_barras_sena_material = codigo_sena_elemento, encargado_registrar_material = encargado_registro, fecha_ingreso_material = '', actualizacion_material = '')

                #Guardar los valores de la variable en el modelo
                agregar_material.save()
        

            return render (request, "App_SistemaGestionInventario/instructor_contratista/registrar_elementos_contrato.html", {'usuarios': usuarios, 'ins_planta':ins_planta, 'tipo_material':material, 'ubicacion':ubicacion, 'estado_material':estado_para_elemento})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/registrar_elementos_contrato.html", {'usuarios': usuarios, 'ins_planta':ins_planta, 'tipo_material':material, 'ubicacion':ubicacion, 'estado_material':estado_para_elemento})



def listar_usuarios_contrato(request):
    #Traer los usuarios y separarlos segun su rol
    ins_planta = Usuario.objects.filter(id_rol = rol[0][0])
    ins_planta = conversion_rol_cuenta(ins_planta)

    ins_contrato = Usuario.objects.filter(id_rol = rol[1][0])
    ins_contrato = conversion_rol_cuenta(ins_contrato)

    monitor = Usuario.objects.filter(id_rol = rol[2][0])
    monitor = conversion_rol_cuenta(monitor)


    #Cambio para que no aparezca una letra y pasar a ver "Activo" o "Inactivo"

    ins_planta = conversion_estado_cuenta(ins_planta)
    ins_contrato = conversion_estado_cuenta(ins_contrato)
    monitor = conversion_estado_cuenta(monitor)

    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_usuarios_contra.html", {'ins_planta':ins_planta, 'ins_contrato':ins_contrato, 'monitor':monitor})



def calendario_contrato(request):
    return render (request, "App_SistemaGestionInventario/instructor_contratista/calendario_contrato.html")



def listar_material_consumible_contrato(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', nombre_material__icontains = busqueda)

        consulta_db = conversion_estado_listar_material(consulta_db)
        
        return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_consu_contrato.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_consu_contrato.html")
    

def listar_material_devolutivo_contrato(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', nombre_material__icontains = busqueda)


        consulta_db = conversion_estado_listar_material(consulta_db)

        
        consulta_db = conversion_estado_listar_material(consulta_db)
        
        return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_devo_contrato.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_devo_contrato.html")



def listar_material_garantia_contrato(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Gara', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Gara', nombre_material__icontains = busqueda)
        
        consulta_db = conversion_estado_listar_material(consulta_db)
        
        return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_garan_contrato.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_garan_contrato.html")
    

def listar_material_soporte_contrato(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Sop', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'Sop', nombre_material__icontains = busqueda)

        
        consulta_db = conversion_estado_listar_material(consulta_db)
        
        return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_sup_contrato.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_sup_contrato.html")
    



def listar_material_baja_contrato(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'DB', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(estado_material = 'DB', nombre_material__icontains = busqueda)

        
        consulta_db = conversion_estado_listar_material(consulta_db)
        
        return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_baja_contrato.html", {'consulta_db':consulta_db, 'busqueda':busqueda})
    return render (request, "App_SistemaGestionInventario/instructor_contratista/listar_mate_baja_contrato.html")



def entregable_consumible_contrato(request):
    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Consu', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])


    #Verificar el metodo por el cual se esta pasando los datos del front
    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_consumible_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_consumible =request.POST.get('fecha_prestamo')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_consumible,
            ubicacion_material_prestamo,
        ]

        #Verifica si los campos de formulario estan llenos (Va a una funcion)
        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("No hay materiales para entregar")
                return render (request, "App_SistemaGestionInventario/instructor_contratista/entregable_consu_contrato.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 
                
            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                return render (request, "App_SistemaGestionInventario/instructor_contrato/entregable_consu_contrato.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 


            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                guardar_nuevo_cliente.save()
            

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_consumible = PrestamosConsumibles(encargado_registra_prestamo_consumible = encargado_prestamo,
                                                                        recibe_prestamo_prestamo_consumible = numero_documento_recibe,
                                                                        ubicacion_prestamo_prestamo_consumible = ubicacion_material_prestamo,
                                                                        material_otorgado_prestamo_consumible = material_entregable,
                                                                        fecha_entrega_prestamo_consumible = fecha_otorga_consumible)
            registrar_nuevo_prestamo_consumible.save()

            #Actualizar el estado del material entregado

            material_entregable.estado_material = estado_material[5][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
               
        else:
            print("sssssssssssssssssssssssssssssssssssssssssssss")
    
    return render(request, "App_SistemaGestionInventario/instructor_contratista/entregable_consu_contrato.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_consu':materiales, 'lugar':ubicacion})



def entregable_devolutivo_contrato(request):
    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Devo', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])

    
    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('rol_persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_devolutivo_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_devolutivo = request.POST.get('fecha_prestamo')
        fecha_devolucion_devolutivo = request.POST.get('fecha_devolucion')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_devolutivo,
            ubicacion_material_prestamo,
            fecha_devolucion_devolutivo,
        ]

        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("CAMINO 3")
                return render (request, "App_SistemaGestionInventario/instructor_contratista/entregable_devo_contrato.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            print("CAMINO 1")

            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                print("CAMINO 4")
                return render (request, "App_SistemaGestionInventario/instructor_contratista/entregable_devo_contrato.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            
            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                print("CAMINO 5")
                guardar_nuevo_cliente.save()

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_devolutivo = PrestamosDevolutivos(encargado_registra_material_devolutivo = encargado_prestamo,
                                                                        recibe_prestamo_material_devolutivo = numero_documento_recibe,
                                                                        ubicacion_prestamo_material_devolutivo = ubicacion_material_prestamo,
                                                                        material_otorgado_devolutivo = material_entregable,
                                                                        fecha_entrega_material_devolutivo = '',
                                                                        fecha_devolucion_material_devolutivo = fecha_devolucion_devolutivo)
            registrar_nuevo_prestamo_devolutivo.save()

            #Actualizar el estado del material entregado
            material_entregable.estado_material = estado_material[1][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
            print("CAMINO 6")



            fecha_entrega = datetime.strptime(fecha_otorga_devolutivo, '%Y-%m-%d').date()
            fecha_devolucion = datetime.strptime(fecha_devolucion_devolutivo, '%Y-%m-%d').date()
            fecha_actual = datetime.now()

            un_dia_antes = fecha_devolucion - timedelta(days=0)
        
            #Enviar correos electrónicos de recordatorio uno días antes
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)


        else:
            print("CAMINO 2")
    return render(request, "App_SistemaGestionInventario/instructor_contratista/entregable_devo_contrato.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})


#Vistas para el monitor

def funciones_monitor(request):
    return render (request, "App_SistemaGestionInventario/monitor/funciones_monitor.html")



def listar_usuarios_monitor(request):
    #Traer los usuarios y separarlos segun su rol
    ins_planta = Usuario.objects.filter(id_rol = rol[0][0])
    ins_planta = conversion_rol_cuenta(ins_planta)

    ins_contrato = Usuario.objects.filter(id_rol = rol[1][0])
    ins_contrato = conversion_rol_cuenta(ins_contrato)

    monitor = Usuario.objects.filter(id_rol = rol[2][0])
    monitor = conversion_rol_cuenta(monitor)


    #Cambio para que no aparezca una letra y pasar a ver "Activo" o "Inactivo"

    ins_planta = conversion_estado_cuenta(ins_planta)
    ins_contrato = conversion_estado_cuenta(ins_contrato)
    monitor = conversion_estado_cuenta(monitor)
    return render (request, "App_SistemaGestionInventario/monitor/listar_usuarios_monitor.html", {'ins_planta':ins_planta, 'ins_contrato':ins_contrato, 'monitor':monitor})

def listar_material_devolutivo_monitor(request):
    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if  metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Devo', nombre_material__icontains = busqueda)



        consulta_db = conversion_estado_listar_material(consulta_db)
        

        materiales_disponibles = []

        for consulta in consulta_db:
            if consulta.estado_material == 'Disponible':
                materiales_disponibles.append(consulta)
        cantidad_modelo = []


        for consulta in materiales_disponibles:
            modelo_actual = consulta.modelo_material
            encontrado = False

            for lista in cantidad_modelo:
                if lista[0] == modelo_actual:
                    lista[1] += 1
                    encontrado = True
                    break

            if not encontrado:
                cantidad_modelo.append([modelo_actual, 1])
        
        #cantidad[] = Materiales.objects.filter(tipo_material = 'Devo', modelo = (Modelo))

        return render (request, "App_SistemaGestionInventario/monitor/listar_mate_devo_monitor.html", {'consulta_db': consulta_db, 'busqueda':busqueda, 'cantidad_modelo':cantidad_modelo})
    return render (request, "App_SistemaGestionInventario/monitor/listar_mate_devo_monitor.html")


def listar_material_consumible_monitor(request):

    if request.method == 'POST':
        metodo_busqueda = request.POST.get('buscar_material')
        busqueda = ''
        consulta_db = ''

        if metodo_busqueda == 'codigo_sena_material':
            busqueda = request.POST.get('CodigoSenaMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', codigo_barras_sena_material__icontains = busqueda)

        elif metodo_busqueda == 'nombre_material':
            busqueda = request.POST.get('NombreMaterial')
            consulta_db = Materiales.objects.filter(tipo_material = 'Consu', nombre_material__icontains = busqueda)


        consulta_db = conversion_estado_listar_material(consulta_db)


        materiales_disponibles = []

        for consulta in consulta_db:
            if consulta.estado_material == 'Disponible':
                materiales_disponibles.append(consulta)


        cantidad_modelo = []


        for consulta in materiales_disponibles:
            modelo_actual = consulta.modelo_material
            encontrado = False

            for lista in cantidad_modelo:
                if lista[0] == modelo_actual:
                    lista[1] += 1
                    encontrado = True
                    break

            if not encontrado:
                cantidad_modelo.append([modelo_actual, 1])

        return render (request, "App_SistemaGestionInventario/monitor/listar_mate_consu_monitor.html", {'materiales_consumibles':consulta_db, 'busqueda':busqueda, 'cantidad_modelo':cantidad_modelo})
    return render (request, "App_SistemaGestionInventario/monitor/listar_mate_consu_monitor.html")

def calendario_monitor(request):
    return render (request, "App_SistemaGestionInventario/monitor/calendario_monitor.html")



def registrar_material_monitor(request):
    #Bloque para desplegables para registrar materiales
    usuarios = Usuario.objects.all()

    #Instructores que se asignan materiales (planta)
    ins_planta = Usuario.objects.filter(id_rol = 'I.P')
    
    lugares = []
    for lugar in ubicacion_material:
        lugares.append(lugar[1])
        
    
    material_estado = []
    for estado in estado_material:
        material_estado.append(estado[1])
        

    material_tipo = []
    for tipo in tipo_material:
        material_tipo.append(tipo[1])


    
    #Bloque para confirmacion e ingreso de nuevos materiales al sistema
    if request.method == 'POST':
        
        #Traer los valores del front
        tipo_elemento = request.POST.get('tipo_material')
        nombre_elemento = request.POST.get('nombre_material')
        modelo_elemento = request.POST.get('modelo_material')
        ubicacion_elemento = request.POST.get('ubicacion_material')
        valor_elemento = request.POST.get('valor_material')
        estado_elemento = request.POST.get('estado_material')
        especificacion_elemento = request.POST.get('especificacion_material')
        instructor_encargado_elemento = request.POST.get('instructor_encargado')
        codigo_barras_normal_elemento = str(request.POST.get('codigo_barras_normal'))
        codigo_sena_elemento = str(request.POST.get('codigo_sena'))
        encargado_registro = request.POST.get('encargado_registro')
        firma_electronica_elemento = request.POST.get('firma_electronica')
        fecha_ingreso_elemento = request.POST.get('fecha_ingreso')
        
        #Se utiliza para guardar las variables que traen los datos de front
        formulario = [
            tipo_elemento,
            nombre_elemento,
            modelo_elemento,
            ubicacion_elemento,
            valor_elemento,
            estado_elemento,
            especificacion_elemento,
            instructor_encargado_elemento,
            codigo_barras_normal_elemento,
            codigo_sena_elemento,
            encargado_registro,
            firma_electronica_elemento,
            fecha_ingreso_elemento,
        ]
        
        mensaje = ''
        
        #Verifica si todas las variables de la lista contienen datos
        if campos_llenos(formulario) == True:

            #Verifica si el codigo SENA que actualmente trae de front se encuentra o no en la db
            if verificar_codigo_sena(codigo_sena_elemento):

                #Llaves foraneas 
                instructor_encargado_elemento = traer_objeto_usuario(instructor_encargado_elemento)
                encargado_registro = traer_objeto_usuario(encargado_registro)


                #Valores de los choices

                #Choices para tipo de material
                tipo_elemento = conversion_tipo_material(tipo_elemento)

                #Choices para ubicacion del material
                ubicacion_elemento = conversion_ubicacion_material(ubicacion_elemento)

                #Choices para estado material
                estado_elemento = conversion_estado_material(estado_elemento)
            
                #Crear variable para guardar todos los datos del modelo "Materiales"
                agregar_material = Materiales(tipo_material = tipo_elemento, nombre_material = nombre_elemento, modelo_material = modelo_elemento, ubicacion_material = ubicacion_elemento, valor_material = valor_elemento, estado_material = estado_elemento, especificacion_tecnica_material = especificacion_elemento, instructor_ecargado_material = instructor_encargado_elemento, codigo_barras_original_material = codigo_barras_normal_elemento, codigo_barras_sena_material = codigo_sena_elemento, encargado_registrar_material = encargado_registro, fecha_ingreso_material = '', actualizacion_material = '')
                
                #Guardar los valores de la variable en el modelo
                agregar_material.save()


                mensaje = 'El material fue registrado.'

            else:
                mensaje = 'El código ya existe, vuelva a intentarlo.'
            
        else:
            mensaje = 'Por favor complete todos los campos.'

        return render (request, "App_SistemaGestionInventario/monitor/registrar_elementos_monitor.html", {'usuarios':usuarios, 'ins_planta':ins_planta, 'ubicaciones':lugares, 'estado_material':material_estado, 'tipo_material':material_tipo})
        
    return render (request, "App_SistemaGestionInventario/monitor/registrar_elementos_monitor.html", {'usuarios':usuarios, 'ins_planta':ins_planta, 'ubicaciones':lugares, 'estado_material':material_estado, 'tipo_material':material_tipo})



def entregable_devolutivo_monitor(request):
    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Devo', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])

    
    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('rol_persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_devolutivo_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_devolutivo = request.POST.get('fecha_prestamo')
        fecha_devolucion_devolutivo = request.POST.get('fecha_devolucion')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_devolutivo,
            ubicacion_material_prestamo,
            fecha_devolucion_devolutivo,
        ]

        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("CAMINO 3")
                return render (request, "App_SistemaGestionInventario/monitor/entregable_devo_monitor.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            print("CAMINO 1")

            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                print("CAMINO 4")
                return render (request, "App_SistemaGestionInventario/monitor/entregable_devo_monitor.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})
            
            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                print("CAMINO 5")
                guardar_nuevo_cliente.save()

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_devolutivo = PrestamosDevolutivos(encargado_registra_material_devolutivo = encargado_prestamo,
                                                                        recibe_prestamo_material_devolutivo = numero_documento_recibe,
                                                                        ubicacion_prestamo_material_devolutivo = ubicacion_material_prestamo,
                                                                        material_otorgado_devolutivo = material_entregable,
                                                                        fecha_entrega_material_devolutivo = '',
                                                                        fecha_devolucion_material_devolutivo = fecha_devolucion_devolutivo)
            registrar_nuevo_prestamo_devolutivo.save()

            #Actualizar el estado del material entregado
            material_entregable.estado_material = estado_material[1][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
            print("CAMINO 6")



            fecha_entrega = datetime.strptime(fecha_otorga_devolutivo, '%Y-%m-%d').date()
            fecha_devolucion = datetime.strptime(fecha_devolucion_devolutivo, '%Y-%m-%d').date()
            fecha_actual = datetime.now()

            un_dia_antes = fecha_devolucion - timedelta(days=0)
        
            #Enviar correos electrónicos de recordatorio uno días antes
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)
            enviar_recordatorio_correo(encargado_prestamo, numero_documento_recibe, material_entregable, fecha_entrega, un_dia_antes)


        else:
            print("CAMINO 2")
    return render (request, "App_SistemaGestionInventario/monitor/entregable_devo_monitor.html", {'encargado_prestar':prestamistas, 'rol_recibe':roles, 'documento':documento_tipo, 'material_devo':materiales, 'lugar':ubicacion})


def entregable_consumible_monitor(request):
    #Datos para desplegables

    #Usuarios que usan el sistema
    prestamistas = Usuario.objects.all()

    #Choices de Rol de persona que recibe material
    roles = []
    for rol in recibe_material:
        roles.append(rol[1])


    #Choices de los tipos de documentos
    documento_tipo = []
    for tipo in tipo_documento:
        documento_tipo.append(tipo[1])
    

    #Materiales consumibles disponibles
    materiales = Materiales.objects.filter(tipo_material = 'Consu', estado_material = 'Dis')
    

    #Donde se otorga el material
    ubicacion = []
    for lugar in ubicacion_material:
        ubicacion.append(lugar[1])


    #Verificar el metodo por el cual se esta pasando los datos del front
    if request.method == 'POST':
        numero_ficha_cliente = ''

        #Traer los datos de los campos de front
        encargado_prestamo = request.POST.get('encargado')
        rol_persona_recibe = request.POST.get('persona_recibe')
        tipo_documento_recibe = request.POST.get('tipo_documento_recibe')
        numero_documento_recibe = request.POST.get('numero_documento_recibe')
        primer_nombre_recibe = request.POST.get('primer_nombre_recibe')
        segundo_nombre_recibe = request.POST.get('segundo_nombre_recibe')
        primer_apellido_recibe = request.POST.get('primer_apellido_recibe')
        segundo_apellido_recibe = request.POST.get('segundo_apellido_recibe')
        correo_sena_persona_recibe = request.POST.get('correo_persona_recibe')
        primer_telefono = request.POST.get('primer_telefono')
        segundo_telefono = request.POST.get('segundo_telefono')
        numero_ficha_cliente = request.POST.get('numero_ficha')
        material_entregable = request.POST.get('material_consumible_entregable')
        ubicacion_material_prestamo = request.POST.get('ubicacion')
        fecha_otorga_consumible =request.POST.get('fecha_prestamo')

        #Lista donde se guardan los datos mas relevantes
        formulario = [
            encargado_prestamo,
            rol_persona_recibe,
            tipo_documento_recibe,
            numero_documento_recibe,
            primer_nombre_recibe,
            primer_apellido_recibe,
            correo_sena_persona_recibe,
            primer_telefono,
            material_entregable,
            fecha_otorga_consumible,
            ubicacion_material_prestamo,
        ]

        #Verifica si los campos de formulario estan llenos (Va a una funcion)
        if campos_llenos(formulario) == True:
            #Cmprobar el valor que trae codigo de barras
            if material_entregable == '-1':
                print("No hay materiales para entregar")
                return render (request, "App_SistemaGestionInventario/monitor/entregable_consu_monitor.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 
                
            #Verificar que si la persona del prestamo es aprendiz y si ingreso el numero de la ficha a la que pertenece
            if rol_persona_recibe == recibe_material[2][1] and (numero_ficha_cliente == None or numero_ficha_cliente == ''):
                '''MENSAJE PARA MOSTRAR EN FRONT'''
                return render (request, "App_SistemaGestionInventario/monitor/entregable_consu_monitor.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion}) 


            #Verificar si el cliente que desea el prestamo ya esta en el sistema (Para ponerlo la llave foranea)
            if comprobar_usuario(numero_documento_recibe) == False:
                #Se crea una variable para almacenar todos los datos del formulario para pasarlos a la db
                guardar_nuevo_cliente = Clientes(rol = rol_persona_recibe, tipo_documento = tipo_documento_recibe, numero_documento = numero_documento_recibe, primer_nombre = primer_nombre_recibe, segundo_nombre = segundo_nombre_recibe, primer_apellido = primer_apellido_recibe, segundo_apellido = segundo_apellido_recibe, correo_soy_sena = correo_sena_persona_recibe, primer_telefono = primer_telefono, segundo_telefono = segundo_telefono, numero_ficha = numero_ficha_cliente, fecha_ingres_sistema = '')
                #Guarda los datos en la base de datos
                guardar_nuevo_cliente.save()
            

            #Traer los objetos de las db para las (FK)
            #FK de encargado que realiza el prestamo consumible
            encargado_prestamo = Usuario.objects.get(numero_documento = encargado_prestamo)

            #FK de persona que recibe el prestamo consumible
            numero_documento_recibe = Clientes.objects.get(numero_documento = numero_documento_recibe)

            #FK del material consumibles que se va a entregar
            material_entregable = Materiales.objects.get(codigo_barras_sena_material = material_entregable)


            #Variable para registrar el prestamo consumible
            registrar_nuevo_prestamo_consumible = PrestamosConsumibles(encargado_registra_prestamo_consumible = encargado_prestamo,
                                                                        recibe_prestamo_prestamo_consumible = numero_documento_recibe,
                                                                        ubicacion_prestamo_prestamo_consumible = ubicacion_material_prestamo,
                                                                        material_otorgado_prestamo_consumible = material_entregable,
                                                                        fecha_entrega_prestamo_consumible = fecha_otorga_consumible)
            registrar_nuevo_prestamo_consumible.save()

            #Actualizar el estado del material entregado

            material_entregable.estado_material = estado_material[5][0]
            #Guardar actualizacion del estado del material entregado
            material_entregable.save()
               
        else:
            print("sssssssssssssssssssssssssssssssssssssssssssss")

    return render (request, "App_SistemaGestionInventario/monitor/entregable_consu_monitor.html", {'prestamistas':prestamistas, 'roles':roles, 'documento_tipo':documento_tipo, 'materiales_disponibles':materiales, 'ubicacion':ubicacion})